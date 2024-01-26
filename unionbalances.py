import os
from openpyxl import load_workbook
import getpass
import datetime

# load import workbook
wb = load_workbook(filename="template.xlsx")
balances = wb.active


# load savings workbook
def loadSavings(savings):
    savingsWb = load_workbook(savings)
    return savingsWb.active


# load grant workbook
def loadGrant(grant):
    grantWb = load_workbook(grant)
    return grantWb.active


# load alumni workbook
def loadAlumni(alumni):
    alumniWb = load_workbook(alumni)
    return alumniWb.active

# returns start and end of savings, grant and ring
def findAccountIndexes(accountId, start, balances):
    i = start
    startNewAcc = start  
    while  accountId not in balances['A' + str(i)].value:
        i+=1
        startNewAcc += 1
    return startNewAcc

# columns of relevant info for loops
balanceCol = "G"  # for import sheet
xeroCol = "B"  # for xero sheet
balanceCode = "B"  # for import sheet
xeroCode = "A"  # for xero sheet

# rows on which the respective sections start
startSavings = 2
startGrant = findAccountIndexes("2112", startSavings, balances)
startRing = findAccountIndexes("2114",startGrant, balances)
endRing = len(balances['A'])

print(startSavings, startGrant, startRing)

def importBalances(sheetType, startAccount, endAccount, maxXeroRow):
    n = 8  # starting row in xero file
    for i in range(startAccount, endAccount):  # balanceCode rows
        for j in range(n, maxXeroRow):  # savingsCode rows
            # get society name and code to compare them later
            societyCode = balances[balanceCode + str(i)].value
            societyName = sheetType[xeroCode + str(j)].value

            try:
                if societyCode[-3:] in societyName:
                    # if the name in the xero sheet contains the society code in the template, add balance to template
                    balances[balanceCol + str(i)].value = sheetType[xeroCol + str(j)].value
                    n = j + 1
                    break
                elif int(societyName[:3]) > int(societyCode[-3:]):
                    break
            except:
                break

# saves updated import file with current date in the name
def save_import():
    current_date = datetime.datetime.now()
    date = datetime.datetime.strftime(current_date, '%d%m%Y')
    path = 'C:\\Users\\' + getpass.getuser() + '\\Downloads\\import' + date + '.xlsx'
    wb.save(path)

# import all xero sheets at once
def importAll(savings, grant, alumni):
    sSheet = loadSavings(savings)
    gSheet = loadGrant(grant)
    aSheet = loadAlumni(alumni)

    maxSavingRow = len(sSheet['A']) - 2
    maxGrantRow = len(gSheet['A']) - 2
    maxRingRow = len(aSheet['A']) - 2

    for i in range(startSavings, maxSavingRow):
        print(balances[balanceCode + str(i)].value)

    for i in range(startSavings, maxSavingRow):
        print(sSheet[xeroCode + str(i)].value)

    importBalances(sSheet, startSavings, startGrant, maxSavingRow)  # for savings
    importBalances(gSheet, startGrant, startRing, maxGrantRow)  # for grant
    importBalances(aSheet, startRing, endRing, maxRingRow)  # for ring fenced
    save_import()



